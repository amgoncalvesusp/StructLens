"""Reproducible tabular exports for scientific results."""

from __future__ import annotations

import csv
import json
from dataclasses import asdict
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from PIL import Image

from structlens.core.difference_maps import DistanceDifferenceMatrix
from structlens.core.evidence import EvidenceCard
from structlens.core.interactions import InteractionDifference
from structlens.core.models import AnalysisResult
from structlens.core.msa import MSAColumn
from structlens.core.sites import SiteMetrics


def export_analysis_xlsx(result: AnalysisResult, path: str | Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    if summary is None:
        raise RuntimeError("Workbook did not create a summary sheet")
    summary.title = "Summary"
    summary.append(["Metric", "Value", "Units"])
    summary.append(["Sequence identity", result.sequence_identity, "fraction"])
    summary.append(["Sequence coverage", result.sequence_coverage, "fraction"])
    summary.append(["Strict Cα RMSD", result.strict_rmsd_angstrom, "Å"])
    summary.append(["Refined Cα RMSD", result.refined_rmsd_angstrom, "Å"])
    summary.append(["Mapped residues", result.mapped_residue_count, "residues"])
    summary.append(["Mutation count", result.mutation_count, "events"])
    summary.append(["Alignment decision", result.alignment_decision, ""])
    for cell in summary[1]:
        cell.font = Font(bold=True)

    residues = workbook.create_sheet("Residues")
    residues.append(
        [
            "Reference residue",
            "Target residue",
            "Status",
            "Mutation notation",
            "Cα displacement (Å)",
            "Backbone RMSD (Å)",
            "Side-chain RMSD (Å)",
            "All-heavy-atom RMSD (Å)",
            "Outlier",
            "Key residue",
        ]
    )
    by_index = {
        event.alignment_index: event.canonical_notation for event in result.mutations
    }
    for item in result.correspondences:
        residues.append(
            [
                _residue_label(item.reference),
                _residue_label(item.target),
                item.status.value,
                by_index.get(item.alignment_index, ""),
                item.ca_displacement_angstrom,
                item.backbone_rmsd_angstrom,
                item.sidechain_rmsd_angstrom,
                item.all_heavy_atom_rmsd_angstrom,
                item.is_outlier,
                item.is_key_residue,
            ]
        )

    mutations = workbook.create_sheet("Mutations")
    mutations.append(
        [
            "Alignment index",
            "Kind",
            "Reference",
            "Target",
            "Notation",
            "BLOSUM62",
            "Grantham distance",
            "Physicochemical class",
        ]
    )
    for event in result.mutations:
        mutations.append(
            [
                event.alignment_index,
                event.kind.value,
                _residue_label(event.reference),
                _residue_label(event.target),
                event.canonical_notation,
                event.blosum62_score,
                event.grantham_distance,
                event.physicochemical_class,
            ]
        )
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def export_analysis_csv(result: AnalysisResult, path: str | Path) -> None:
    rows = [
        {
            "reference_residue": _residue_label(item.reference),
            "target_residue": _residue_label(item.target),
            "status": item.status.value,
            "ca_displacement_angstrom": item.ca_displacement_angstrom,
            "backbone_rmsd_angstrom": item.backbone_rmsd_angstrom,
            "outlier": item.is_outlier,
        }
        for item in result.correspondences
    ]
    with Path(path).open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=list(rows[0]) if rows else ["status"]
        )
        writer.writeheader()
        writer.writerows(rows)


def export_analysis_json(result: AnalysisResult, path: str | Path) -> None:
    payload: dict[str, Any] = {
        "reference_id": result.reference_id,
        "target_id": result.target_id,
        "sequence_identity": result.sequence_identity,
        "sequence_coverage": result.sequence_coverage,
        "alignment_decision": result.alignment_decision,
        "strict_rmsd_angstrom": result.strict_rmsd_angstrom,
        "refined_rmsd_angstrom": result.refined_rmsd_angstrom,
        "mutations": [asdict(event) for event in result.mutations],
    }
    Path(path).write_text(
        json.dumps(payload, indent=2, default=str) + "\n", encoding="utf-8"
    )


def export_publication_image(
    image: Image.Image,
    path: str | Path,
    *,
    width_mm: float = 100.0,
    dpi: int = 300,
) -> None:
    """Render a supplied molecular scene at publication dimensions and DPI."""

    if dpi not in {300, 600}:
        raise ValueError("Publication image dpi must be 300 or 600")
    if width_mm <= 0:
        raise ValueError("width_mm must be positive")
    output_path = Path(path)
    suffix = output_path.suffix.lower()
    if suffix not in {".jpg", ".jpeg", ".tif", ".tiff"}:
        raise ValueError("Publication image format must be JPEG or TIFF")
    if suffix in {".jpg", ".jpeg"} and (
        "A" in image.getbands() or image.mode in {"RGBA", "LA"}
    ):
        raise ValueError(
            "JPEG export does not support transparency; provide an opaque image"
        )
    width_px = round(width_mm / 25.4 * dpi)
    height_px = max(1, round(width_px * image.height / image.width))
    rendered = image.resize((width_px, height_px), Image.Resampling.LANCZOS)
    if suffix in {".jpg", ".jpeg"}:
        rendered = rendered.convert("RGB")
        rendered.save(
            output_path, format="JPEG", quality=95, dpi=(dpi, dpi), optimize=True
        )
    else:
        rendered.save(
            output_path, format="TIFF", compression="tiff_deflate", dpi=(dpi, dpi)
        )


def export_v03_xlsx(
    path: str | Path,
    *,
    msa_columns: tuple[MSAColumn, ...] = (),
    interaction_differences: tuple[InteractionDifference, ...] = (),
    site_metrics: tuple[SiteMetrics, ...] = (),
    distance_matrix: DistanceDifferenceMatrix | None = None,
    evidence_cards: tuple[EvidenceCard, ...] = (),
    provenance: tuple[str, ...] = (),
) -> None:
    """Export v0.3 scientific records without recalculating any metric."""
    workbook = Workbook()
    default = workbook.active
    if default is None:
        raise RuntimeError("Workbook did not create a worksheet")
    workbook.remove(default)

    msa = workbook.create_sheet("MSA")
    msa.append(["Alignment column", "Reference label", "Reference residue", "Non-gap count", "Gap fraction", "Ambiguous fraction", "Conservation", "Entropy (bits)"])
    for column in msa_columns:
        msa.append([column.index, column.reference_label, _residue_label(column.reference_residue), column.non_gap_count, column.gap_fraction, column.ambiguous_fraction, column.conservation_score, column.entropy_bits])

    conservation = workbook.create_sheet("Conservation")
    conservation.append(["Alignment column", "Reference label", "Conservation", "Gap fraction", "Ambiguous fraction", "Units"])
    for column in msa_columns:
        conservation.append([column.index, column.reference_label, column.conservation_score, column.gap_fraction, column.ambiguous_fraction, "fraction"])

    frequencies = workbook.create_sheet("Amino Acid Frequencies")
    frequencies.append(["Alignment column", "Reference label", *tuple("ACDEFGHIKLMNPQRSTVWY")])
    for column in msa_columns:
        counts = {letter: 0 for letter in "ACDEFGHIKLMNPQRSTVWY"}
        valid = 0
        for cell in column.cells:
            if cell.character.upper() in counts:
                counts[cell.character.upper()] += 1
                valid += 1
        frequencies.append([column.index, column.reference_label, *(counts[letter] / valid if valid else None for letter in counts)])

    interactions = workbook.create_sheet("Interaction Differences")
    interactions.append(["Type", "Reference position A", "Reference position B", "Change", "Reference distance (Å)", "Target distance (Å)", "Evidence mode"])
    for difference in interaction_differences:
        record = difference.target_record if difference.target_record is not None else difference.reference_record
        interactions.append([difference.key.interaction_type.value, difference.key.reference_position_a, difference.key.reference_position_b, difference.change.value, difference.reference_record.distance_angstrom if difference.reference_record else None, difference.target_record.distance_angstrom if difference.target_record else None, record.evidence_mode if record is not None else None])

    sites = workbook.create_sheet("Sites")
    sites.append(["Site", "Structure", "Mapped residues", "Coverage", "Global-frame RMSD (Å)", "Site-fitted RMSD (Å)", "SASA (Å²)", "Atomic envelope volume (Å³)"])
    for metric in site_metrics:
        sites.append([metric.site_id, metric.structure_id, metric.mapped_residue_count, metric.coverage_fraction, metric.global_frame_backbone_rmsd_angstrom, metric.site_fitted_backbone_rmsd_angstrom, metric.sasa_angstrom2, metric.atomic_envelope_volume_angstrom3])

    fingerprints = workbook.create_sheet("Site Metrics")
    fingerprints.append(["Site", "Structure", "Centroid displacement (Å)", "Radius of gyration (Å)", "Polar fraction", "Charged fraction"])
    for metric in site_metrics:
        fingerprints.append([metric.site_id, metric.structure_id, metric.centroid_displacement_angstrom, metric.radius_of_gyration_angstrom, metric.polar_residue_fraction, metric.charged_residue_fraction])

    if distance_matrix is not None:
        distances = workbook.create_sheet("Distance Difference Matrix")
        delta = np.asarray(distance_matrix.delta_angstrom, dtype=np.float64)
        valid_mask = np.asarray(distance_matrix.valid_mask, dtype=bool)
        distances.append(["Reference position", *distance_matrix.reference_positions])
        for index, label in enumerate(distance_matrix.reference_positions):
            distances.append([label, *[float(value) if valid_mask[index, column] else None for column, value in enumerate(delta[index])]])

    evidence = workbook.create_sheet("Residue Evidence")
    evidence.append(["Reference residue", "Target", "Evidence quality", "Sequence conservation", "Cα displacement (Å)", "Global site RMSD (Å)", "Site-fitted RMSD (Å)"])
    for card in evidence_cards:
        metrics = card.site.metrics[0] if card.site.metrics else None
        evidence.append([_residue_label(card.reference_residue), card.target_id, card.quality.overall_status, card.sequence.conservation_fraction, card.structure.ca_displacement_angstrom, metrics.global_frame_backbone_rmsd_angstrom if metrics else None, metrics.site_fitted_backbone_rmsd_angstrom if metrics else None])

    quality = workbook.create_sheet("Evidence Quality")
    quality.append(["Reference residue", "Target", "Status", "Available sections", "Unavailable sections", "Warnings"])
    for card in evidence_cards:
        quality.append([_residue_label(card.reference_residue), card.target_id, card.quality.overall_status, ", ".join(card.quality.available_sections), ", ".join(card.quality.unavailable_sections), "; ".join(card.quality.warnings)])

    provenance_sheet = workbook.create_sheet("Provenance")
    provenance_sheet.append(["Source"])
    for item in provenance:
        provenance_sheet.append([item])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def _residue_label(residue: Any) -> str:
    if residue is None:
        return ""
    insertion = residue.insertion_code or ""
    return f"{residue.chain_id}:{residue.auth_seq_id}{insertion} {residue.residue_name}"


__all__ = [
    "export_analysis_csv",
    "export_analysis_json",
    "export_analysis_xlsx",
    "export_publication_image",
    "export_v03_xlsx",
]
