from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from structlens.application.analysis_service import AnalysisService
from structlens.application.export_service import export_analysis_xlsx
from tests.unit.application.test_analysis_service import _structure


def test_xlsx_export_contains_units_and_numeric_metrics(tmp_path: Path) -> None:
    result = AnalysisService().analyze(_structure("ref"), _structure("target"))
    output = tmp_path / "analysis.xlsx"

    export_analysis_xlsx(result, output)
    workbook = load_workbook(output, data_only=True)

    assert workbook.sheetnames == ["Summary", "Residues", "Mutations"]
    assert workbook["Summary"]["B2"].value == 1.0
    assert workbook["Residues"]["E1"].value == "Cα displacement (Å)"
