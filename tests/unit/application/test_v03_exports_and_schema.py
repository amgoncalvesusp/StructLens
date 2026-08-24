from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from structlens.application.export_service import export_v03_xlsx
from structlens.application.project_state import ProjectState
from structlens.core.evidence import EvidenceCard, quality_for_sections
from structlens.core.models import ResidueId
from structlens.core.msa import MSAColumn, MSAResidueCell, SequenceResidueRef


def test_v03_project_schema_round_trip_preserves_authoritative_metadata() -> None:
    state = ProjectState(
        msa_settings={"algorithm": "muscle5"},
        interaction_thresholds={"hbond_distance_angstrom": 3.5},
        site_definitions=({"site_id": "active", "mode": "key_residues"},),
        distance_difference_metadata={"valid_pair_count": 4},
        evidence_sources={"A:10": ["msa", "structure"]},
    )
    restored = ProjectState.from_json(state.to_json())
    assert restored.schema_version == "3.0"
    assert restored.msa_settings["algorithm"] == "muscle5"
    assert restored.site_definitions[0]["site_id"] == "active"


def test_v03_xlsx_has_required_sheets_and_blank_unavailable(tmp_path: Path) -> None:
    path = tmp_path / "results.xlsx"
    card = EvidenceCard(ResidueId("ref", "1", "A", "10", None, "ALA"), "target", quality=quality_for_sections(["sequence"], ["structure"]))
    export_v03_xlsx(path, evidence_cards=(card,), provenance=("fixture",))
    workbook = load_workbook(path)
    assert {
        "MSA",
        "Conservation",
        "Amino Acid Frequencies",
        "Insertions",
        "Interaction Differences",
        "Sites",
        "Site Metrics",
        "Site Interaction Fingerprints",
        "Residue Evidence",
        "Evidence Quality",
    }.issubset(workbook.sheetnames)
    assert workbook["Residue Evidence"]["E2"].value is None


def test_v03_xlsx_exports_insertions_and_site_interaction_fingerprints(
    tmp_path: Path,
) -> None:
    target_residue = ResidueId("target", "1", "A", "164", None, "GLY")
    insertion = MSAColumn(
        index=2,
        reference_label="A:163+1",
        reference_residue=None,
        cells=(
            MSAResidueCell(
                "target",
                2,
                SequenceResidueRef(163, "G", target_residue),
                "G",
            ),
        ),
        non_gap_count=1,
        gap_fraction=0.0,
        ambiguous_fraction=0.0,
        conservation_score=None,
        entropy_bits=None,
    )
    path = tmp_path / "v03.xlsx"

    export_v03_xlsx(
        path,
        msa_columns=(insertion,),
        site_interaction_fingerprints=(
            {
                "site_id": "active",
                "structure_id": "target",
                "interaction_type": "hbond_geometric",
                "reference_count": 2,
                "target_count": 1,
                "conserved_count": 1,
                "gained_count": 0,
                "lost_count": 1,
            },
        ),
    )

    workbook = load_workbook(path)
    insertions = workbook["Insertions"]
    assert insertions["A2"].value == 2
    assert insertions["B2"].value == "A:163+1"
    assert insertions["C2"].value == "target"
    assert insertions["D2"].value == 163
    assert insertions["E2"].value == "G"
    assert insertions["F2"].value == "A:164 GLY"

    fingerprints = workbook["Site Interaction Fingerprints"]
    assert fingerprints["A2"].value == "active"
    assert fingerprints["C2"].value == "hbond_geometric"
    assert fingerprints["D2"].value == 2
    assert fingerprints["H2"].value == 1


def test_v03_xlsx_uses_blank_cells_for_unavailable_fingerprint_values(
    tmp_path: Path,
) -> None:
    path = tmp_path / "unavailable.xlsx"
    export_v03_xlsx(
        path,
        site_interaction_fingerprints=(
            {
                "site_id": "active",
                "structure_id": "target",
                "interaction_type": "metal_contact",
                "reference_count": None,
                "target_count": None,
                "conserved_count": None,
                "gained_count": None,
                "lost_count": None,
            },
        ),
    )

    fingerprints = load_workbook(path)["Site Interaction Fingerprints"]
    assert fingerprints["D2"].value is None
    assert fingerprints["H2"].value is None
